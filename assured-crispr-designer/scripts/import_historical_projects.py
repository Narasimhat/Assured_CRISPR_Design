from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


WORKBOOK_NAME = "Project3_CRISPR_RNPs_LAGESO_S1_Documentation_Complete.xlsx"
SEARCH_ROOT = Path(r"U:\DATA MANAGMENT\Core organisation\Administration")
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "src" / "data" / "historicalProjects.js"
SHEET_NAME = "13_15 SD Proj 3 CRISPR_RNP"
HEADER_ROW = 3


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def extract_gene_parts(raw: object) -> tuple[str, str]:
    text = clean(raw)
    if not text:
        return "", ""
    match = re.match(r"([^()]+?)\s*\(([^()]+)\)", text)
    if match:
        return clean(match.group(1)), clean(match.group(2))
    return text, ""


def normalize_mod_type(raw: object) -> str:
    value = clean(raw).upper()
    if "SNP" in value or "POINT" in value or "MUT" in value:
        return "pm"
    if "KO" in value or "KNOCKOUT" in value:
        return "ko"
    if "KI" in value or "KNOCKIN" in value or "TAG" in value or "REPORTER" in value:
        return "ki"
    return "other"


def infer_donor_type(raw: object) -> str:
    text = clean(raw).upper()
    if not text:
        return "none"
    if "SSODN" in text or "ODN" in text:
        return "ssODN"
    return "donor"


def parse_guides(raw: object) -> list[dict[str, str]]:
    text = clean(raw)
    if not text:
        return []
    matches = re.findall(r"([A-Za-z0-9_\-]+)\s*:\s*([ACGTacgt ]{18,})", text)
    if matches:
        guides = []
        for name, seq in matches:
            sequence = re.sub(r"[^ACGTacgt]", "", seq).upper()
            if sequence:
                guides.append({"name": clean(name), "sequence": sequence})
        return guides
    sequences = re.findall(r"[ACGTacgt]{18,}", text)
    return [{"name": f"gRNA{index + 1}", "sequence": seq.upper()} for index, seq in enumerate(sequences)]


def parse_donor(raw: object) -> dict[str, str]:
    text = clean(raw)
    if not text:
        return {"name": "", "sequence": ""}
    match = re.match(r"([^:]+):\s*([ACGTacgt ]+)$", text)
    if match:
        return {
            "name": clean(match.group(1)),
            "sequence": re.sub(r"[^ACGTacgt]", "", match.group(2)).upper(),
        }
    return {"name": "", "sequence": "".join(re.findall(r"[ACGTacgt]+", text)).upper()}


def locate_workbook() -> Path:
    try:
        return next(SEARCH_ROOT.rglob(WORKBOOK_NAME))
    except StopIteration as error:
        raise FileNotFoundError(f"Could not find {WORKBOOK_NAME} under {SEARCH_ROOT}") from error


def build_records(workbook_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook[SHEET_NAME]
    headers = [sheet.cell(HEADER_ROW, column).value for column in range(1, sheet.max_column + 1)]
    records: list[dict[str, object]] = []

    for row_index in range(HEADER_ROW + 1, sheet.max_row + 1):
        row = {clean(headers[index]): sheet.cell(row_index, index + 1).value for index in range(len(headers))}
        project_id = clean(row.get("Lfd.Nr."))
        if not project_id:
            continue
        if clean(row.get("Discarded")).lower() in {"yes", "y", "true", "discarded"}:
            continue

        gene, gene_id = extract_gene_parts(row.get("Target Gene (NCBI/ENSEMBL ID)"))
        donor = parse_donor(row.get("Donor Seq / ssODN"))
        records.append({
            "projectId": project_id,
            "parentalLine": clean(row.get("Parental Cell Line / Donor")),
            "species": clean(row.get("Species (Recipient)")),
            "tool": clean(row.get("Vector / Tool")),
            "guideSpecies": clean(row.get("sgRNA Species")),
            "targetGene": gene,
            "targetGeneId": gene_id,
            "donorType": infer_donor_type(row.get("Donor Seq / ssODN")),
            "donorName": donor["name"],
            "donorSequence": donor["sequence"],
            "establishedLine": clean(row.get("Cell Line Name (hPSCreg)")),
            "receivedFrom": clean(row.get("Received From")),
            "modificationType": clean(row.get("Modification Type")),
            "modificationClass": normalize_mod_type(row.get("Modification Type")),
            "modificationDescription": clean(row.get("Modification Description")),
            "zygosity": clean(row.get("Zygosity")),
            "registrationDate": clean(row.get("Date of Registration")),
            "hazardPotential": clean(row.get("Hazard Potential")),
            "diseaseModel": clean(row.get("Disease / Research Model")),
            "guides": parse_guides(row.get("sgRNA / Targeting Sequence")),
            "successStatus": "established",
        })

    summary = {
        "sourceWorkbook": str(workbook_path),
        "sheet": sheet.title,
        "recordCount": len(records),
        "byClass": {},
    }
    for record in records:
        key = record["modificationClass"]
        summary["byClass"][key] = summary["byClass"].get(key, 0) + 1

    return records, summary


def write_module(records: list[dict[str, object]], summary: dict[str, object]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    module = (
        "// Auto-generated by scripts/import_historical_projects.py\n"
        "export const HISTORICAL_PROJECTS_SUMMARY = "
        f"{json.dumps(summary, indent=2)};\n\n"
        "export const HISTORICAL_PROJECTS = "
        f"{json.dumps(records, indent=2)};\n"
    )
    OUTPUT_PATH.write_text(module, encoding="utf-8")


def main() -> None:
    workbook_path = locate_workbook()
    records, summary = build_records(workbook_path)
    write_module(records, summary)
    print(f"Wrote {len(records)} historical projects to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
